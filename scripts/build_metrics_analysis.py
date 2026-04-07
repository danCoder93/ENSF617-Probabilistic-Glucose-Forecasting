from pathlib import Path
import json
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

ROOT = Path("artifacts/main_run")
OUT = ROOT / "analysis_outputs"
OUT.mkdir(parents=True, exist_ok=True)

# Load artifacts
metrics_summary = json.loads((ROOT / "metrics_summary.json").read_text())
run_summary = json.loads((ROOT / "run_summary.json").read_text())
data_summary = json.loads((ROOT / "data_summary.json").read_text())

df_h = pd.read_csv(ROOT / "metrics_by_horizon.csv")
df_s = pd.read_csv(ROOT / "metrics_by_subject.csv")
df_r = pd.read_csv(ROOT / "metrics_by_glucose_range.csv")

# --- Plots ---
plt.figure(figsize=(8, 5))
plt.plot(df_h["group_value"], df_h["mae"], marker="o", label="MAE")
plt.plot(df_h["group_value"], df_h["rmse"], marker="o", label="RMSE")
plt.xlabel("Forecast horizon step (5-minute intervals)")
plt.ylabel("Error (mg/dL)")
plt.title("Error increases with forecast horizon")
plt.legend()
plt.tight_layout()
h1 = OUT / "horizon_error_plot.png"
plt.savefig(h1, dpi=180)
plt.close()

plt.figure(figsize=(8, 5))
x = range(len(df_r))
plt.bar(x, df_r["mae"], width=0.4, label="MAE")
plt.bar([i + 0.4 for i in x], df_r["rmse"], width=0.4, label="RMSE")
plt.xticks([i + 0.2 for i in x], df_r["group_value"])
plt.ylabel("Error (mg/dL)")
plt.title("Performance is strongest in the 70–180 mg/dL range")
plt.legend()
plt.tight_layout()
h2 = OUT / "glucose_range_error_plot.png"
plt.savefig(h2, dpi=180)
plt.close()

df_s_plot = df_s[df_s["count"] >= 1000].sort_values("mae")
plt.figure(figsize=(8, 8))
plt.barh(df_s_plot["group_value"], df_s_plot["mae"])
plt.xlabel("MAE (mg/dL)")
plt.ylabel("Subject")
plt.title("Subject-level variability in MAE")
plt.tight_layout()
h3 = OUT / "subject_mae_plot.png"
plt.savefig(h3, dpi=180)
plt.close()

# --- Executive summary text ---
headline = metrics_summary["test_metrics"][0]
best_subject = df_s_plot.loc[df_s_plot["mae"].idxmin()]
worst_subject = df_s_plot.loc[df_s_plot["mae"].idxmax()]
best_range = df_r.loc[df_r["mae"].idxmin()]
worst_range = df_r.loc[df_r["mae"].idxmax()]
best_h = df_h.loc[df_h["mae"].idxmin()]
worst_h = df_h.loc[df_h["mae"].idxmax()]
coverage_anomaly = df_h.loc[df_h["empirical_interval_coverage"].idxmin()]

summary_md = f"""# Executive Summary: Local 3-Epoch Run

## Bottom line
This run shows the model is learning meaningful signal, but it is not yet ready for strong real-world claims without longer training and better extreme-range reliability.

## Headline metrics
- Test MAE: **{headline['test_mae']:.2f} mg/dL**
- Test RMSE: **{headline['test_rmse']:.2f} mg/dL**
- Test prediction interval width: **{headline['test_prediction_interval_width']:.2f} mg/dL**
- Test target mean: **{headline['test_target_mean']:.2f} mg/dL**
- Predicted mean: **{headline['test_quantile_prediction_mean']:.2f} mg/dL**

## What looks encouraging
- Aggregate prediction and target means are closely aligned.
- The model performs best in the **{best_range['group_value']}** range.
- Error grows with horizon in a smooth, believable way.

## What needs attention
- Worst glucose range is **{worst_range['group_value']}** with MAE **{worst_range['mae']:.2f} mg/dL**.
- Worst horizon is **{int(worst_h['group_value'])}** with MAE **{worst_h['mae']:.2f} mg/dL**.
- Best subject is **{best_subject['group_value']}** with MAE **{best_subject['mae']:.2f} mg/dL**.
- Worst subject is **{worst_subject['group_value']}** with MAE **{worst_subject['mae']:.2f} mg/dL**.
- Coverage anomaly at horizon **{int(coverage_anomaly['group_value'])}**: empirical interval coverage **{coverage_anomaly['empirical_interval_coverage']:.4f}**.

## Interpretation
This is a strong systems-validation run. The stack works end-to-end, the model learns, and the artifacts are useful.
The real issue now is not “does it run?” but “where is it reliable, where is it weak, and how clinically useful is it?”
"""

summary_path = OUT / "executive_summary_metrics_analysis.md"
summary_path.write_text(summary_md, encoding="utf-8")

# --- Workbook ---
wb = Workbook()
header_fill = PatternFill("solid", fgColor="1F4E78")
sub_fill = PatternFill("solid", fgColor="D9EAF7")
thin = Side(style="thin", color="D0D7DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws = wb.active
ws.title = "Executive Summary"
rows = [
    ["Executive summary metric", "Value", "Interpretation"],
    ["Test MAE (mg/dL)", headline["test_mae"], "Average absolute error."],
    ["Test RMSE (mg/dL)", headline["test_rmse"], "Sensitive to large misses."],
    ["Prediction interval width (mg/dL)", headline["test_prediction_interval_width"], "How wide the uncertainty band is."],
    ["Best glucose range", best_range["group_value"], f"MAE {best_range['mae']:.2f} mg/dL"],
    ["Worst glucose range", worst_range["group_value"], f"MAE {worst_range['mae']:.2f} mg/dL"],
    ["Best horizon", int(best_h["group_value"]), f"MAE {best_h['mae']:.2f} mg/dL"],
    ["Worst horizon", int(worst_h["group_value"]), f"MAE {worst_h['mae']:.2f} mg/dL"],
    ["Best subject", best_subject["group_value"], f"MAE {best_subject['mae']:.2f} mg/dL"],
    ["Worst subject", worst_subject["group_value"], f"MAE {worst_subject['mae']:.2f} mg/dL"],
    ["Biggest red flag", f"Horizon {int(coverage_anomaly['group_value'])}", f"Coverage {coverage_anomaly['empirical_interval_coverage']:.4f}"],
]
for row in rows:
    ws.append(row)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(color="FFFFFF", bold=True)
    cell.border = border
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for c in row:
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 50

def add_df_sheet(name, df):
    ws = wb.create_sheet(title=name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border = border

add_df_sheet("Overall Headline", pd.DataFrame(metrics_summary["test_metrics"]))
add_df_sheet("By Horizon", df_h)
add_df_sheet("By Subject", df_s)
add_df_sheet("By Glucose Range", df_r)

ws = wb.create_sheet("Plots")
ws["A1"] = "Figure"
ws["B1"] = "Why it matters"
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(color="FFFFFF", bold=True)
ws["A2"] = "Horizon error plot"
ws["B2"] = "Shows how forecast quality decays as horizon extends."
ws["A22"] = "Glucose range error plot"
ws["B22"] = "Shows in-range vs hypo/hyper performance."
ws["A42"] = "Subject MAE plot"
ws["B42"] = "Shows subject-level heterogeneity."

for img_path, anchor in [(h1, "A3"), (h2, "A23"), (h3, "A43")]:
    img = XLImage(str(img_path))
    img.width = 720
    img.height = 450
    ws.add_image(img, anchor)

xlsx_path = OUT / "metrics_analysis_workbook.xlsx"
wb.save(xlsx_path)

print(f"Created {xlsx_path}")
print(f"Created {summary_path}")
print(f"Created {h1}")
print(f"Created {h2}")
print(f"Created {h3}")